import inspect
import logging
from openpyxl import Workbook, load_workbook


class Utils():
    def __init__(self):
        self.log_level = logging.DEBUG

    def set_log_level(self, log_level):
        self.log_level = log_level

    # Return logger object for logging test information
    def custom_logger(self):
        # Set to class name, helps with troubleshooting when looking at automation log
        logger_name = inspect.stack()[1][3]
        # logger creation
        logger = logging.getLogger(logger_name)
        logger.setLevel(self.log_level)
        # create handler
        file_handler = logging.FileHandler("..\\Reports\\automation.log", mode="a")
        # create formatter and add to handler
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt="%m/%d/%Y %I:%M:%S %p")
        file_handler.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(file_handler)
        return logger

    @staticmethod
    def read_data_from_excel(file_name, sheet):
        data_list = []
        work_book = load_workbook(filename=file_name)
        sheet = work_book[sheet]
        row_ct = sheet.max_row
        col_ct = sheet.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sheet.cell(row=i, column=j).value)
            data_list.append(row)
        print(data_list)
        return data_list
