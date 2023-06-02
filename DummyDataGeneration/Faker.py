from faker import Faker
from faker.providers import lorem
from openpyxl import Workbook

# fake data generator object, use_weighting = whether all frequencies have equal chance or follow real-world
# False is faster
fake_data = Faker(use_weighting=False)
# data generation in Japanese locale
# fake_data = Faker('ja_JP')
# random data between localisation
# fake_data= Faker(['en_US', 'ja_JP'])
# add lorem data generator
fake_data.add_provider(lorem)
# a seed provides same set of data used for unit testing
Faker.seed(2099)
# to switch to seed
# fake_data.seed_instance(2099)
# examples (each call generates a different sample)
print(fake_data.name())
# unique value
print(fake_data.unique.name())
print(fake_data.email())
print(fake_data.address())
print(fake_data.text())

# lorem provider sample usage
# nb_sentences = # of sentences, variable_nb_sentences = varied true or false, ext_word_list = use words from array
print(fake_data.paragraph(nb_sentences=5, variable_nb_sentences=False))
print(fake_data.paragraph(nb_sentences=5, variable_nb_sentences=False, ext_word_list=['cat', 'dog', 'bird', 'duck']))
# lists of paragraphs
print(fake_data.paragraphs(nb=5))
# a sentence
print(fake_data.sentence(nb_words=10))
print(fake_data.sentences(nb=10))
# text string
print(fake_data.text(max_nb_chars=20))
# word
print(fake_data.word(ext_word_list=['cat', 'dog', 'bird', 'duck']))


wb = Workbook()
# worksheet object
ws = wb.active

# rows
for i in range(1, 11):
    # insert 3 column data per row
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()
    ws.cell(row=i, column=3).value = fake_data.address()

# save workbook as specified name
wb.save("fakerData.xlsx")
